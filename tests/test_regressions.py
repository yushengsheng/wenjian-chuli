from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from zipfile import ZipFile
from xml.etree import ElementTree as ET

import pandas as pd

from spreadsheet_tool.comparison import (
    align_for_comparison,
    compare_row_values,
    get_ignored_compare_columns,
    preview_value,
)
from spreadsheet_tool.models import ColumnSetting, ExportSettings, PipelineConfig, SourceSelection
from spreadsheet_tool.processor import (
    INTERNAL_APPEND_ORDER,
    INTERNAL_SOURCE_FILE,
    INTERNAL_SOURCE_ROLE,
    INTERNAL_SOURCE_SHEET,
    build_filter_mask,
    canonical_internal_column_name,
    combine_enabled_sources,
    collect_target_columns,
    default_display_name,
    export_dataframe_with_old_workbook,
    load_sources_from_paths,
    materialize_dataframe,
    process_dataframe,
    read_excel_sheet,
    values_differ,
    write_dataframe_to_existing_excel_sheet,
)


class ProcessorRegressionTests(unittest.TestCase):
    def create_workbook_with_empty_fill_style(self, path: Path) -> Path:
        from io import BytesIO

        from openpyxl import Workbook, load_workbook

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet.append(["email", "note"])
        worksheet.append(["a@example.com", "hello"])
        workbook.save(path)
        workbook.close()

        namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        with ZipFile(path) as source_zip:
            styles_root = ET.fromstring(source_zip.read("xl/styles.xml"))
            fills = styles_root.find(f"{{{namespace}}}fills")
            self.assertIsNotNone(fills)
            assert fills is not None
            fills.append(ET.Element(f"{{{namespace}}}fill"))
            fills.set("count", str(len(list(fills))))

            rebuilt_path = path.with_name(f"{path.stem}_broken{path.suffix}")
            with ZipFile(rebuilt_path, mode="w") as target_zip:
                for item in source_zip.infolist():
                    payload = source_zip.read(item.filename)
                    if item.filename == "xl/styles.xml":
                        payload = ET.tostring(styles_root, encoding="utf-8", xml_declaration=True)
                    target_zip.writestr(item, payload)

        with self.assertRaisesRegex(TypeError, "openpyxl.styles.fills.Fill"):
            load_workbook(BytesIO(rebuilt_path.read_bytes()), read_only=True, data_only=True)

        return rebuilt_path

    def test_values_differ_uses_field_level_comparison_semantics(self) -> None:
        self.assertFalse(values_differ("same@example.com", "same@example.com"))
        self.assertFalse(values_differ(pd.NA, None))
        self.assertFalse(values_differ(1, 1.0))
        self.assertTrue(values_differ("old-pass", "new-pass"))

    def test_load_sources_from_paths_repairs_empty_fill_stylesheet_nodes(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            broken_path = self.create_workbook_with_empty_fill_style(Path(tmp_dir) / "source.xlsx")

            sources, cache = load_sources_from_paths([broken_path])

        self.assertEqual(len(sources), 1)
        source = sources[0]
        self.assertEqual(source.sheet_name, "Sheet1")
        self.assertEqual(cache[source.source_id].to_dict(orient="records"), [{"email": "a@example.com", "note": "hello"}])

    def test_read_excel_sheet_repairs_empty_fill_stylesheet_nodes(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            broken_path = self.create_workbook_with_empty_fill_style(Path(tmp_dir) / "sheet.xlsx")

            dataframe = read_excel_sheet(broken_path, "Sheet1")

        self.assertEqual(dataframe.to_dict(orient="records"), [{"email": "a@example.com", "note": "hello"}])

    def test_materialize_dataframe_uses_first_non_empty_row_as_header_for_single_row_input(self) -> None:
        raw = pd.DataFrame([["a@example.com", "pw123456"]])

        result = materialize_dataframe(raw)

        self.assertEqual(result.columns.tolist(), ["a@example.com", "pw123456"])
        self.assertEqual(result.to_dict(orient="records"), [])

    def test_materialize_dataframe_uses_first_non_empty_row_as_header_for_multiple_rows(self) -> None:
        raw = pd.DataFrame(
            [
                ["a@example.com", "pw123456"],
                ["b@example.com", "pw999999"],
            ]
        )

        result = materialize_dataframe(raw)

        self.assertEqual(result.columns.tolist(), ["a@example.com", "pw123456"])
        self.assertEqual(
            result.to_dict(orient="records"),
            [
                {"a@example.com": "b@example.com", "pw123456": "pw999999"},
            ],
        )

    def test_materialize_dataframe_ignores_leading_empty_rows_before_using_first_header_row(self) -> None:
        raw = pd.DataFrame(
            [
                [pd.NA, pd.NA],
                ["邮箱", "密码"],
                ["a@example.com", "pw123456"],
            ]
        )

        result = materialize_dataframe(raw)

        self.assertEqual(result.columns.tolist(), ["邮箱", "密码"])
        self.assertEqual(result.to_dict(orient="records"), [{"邮箱": "a@example.com", "密码": "pw123456"}])

    def test_materialize_dataframe_ignores_columns_with_blank_first_row_titles(self) -> None:
        raw = pd.DataFrame(
            [
                ["邮箱", pd.NA, "密码"],
                ["a@example.com", "should-be-ignored", "pw123456"],
            ]
        )

        result = materialize_dataframe(raw)

        self.assertEqual(result.columns.tolist(), ["邮箱", "密码"])
        self.assertEqual(result.to_dict(orient="records"), [{"邮箱": "a@example.com", "密码": "pw123456"}])

    def test_materialize_dataframe_normalizes_legacy_internal_display_columns(self) -> None:
        raw = pd.DataFrame(
            [
                [
                    "邮箱",
                    default_display_name(INTERNAL_SOURCE_FILE),
                    default_display_name(INTERNAL_SOURCE_SHEET),
                    default_display_name(INTERNAL_SOURCE_ROLE),
                    default_display_name(INTERNAL_APPEND_ORDER),
                ],
                ["a@example.com", "old.xlsx", "Sheet1", "old", 1],
            ]
        )

        result = materialize_dataframe(raw)

        self.assertEqual(
            result.columns.tolist(),
            ["邮箱", INTERNAL_SOURCE_FILE, INTERNAL_SOURCE_SHEET, INTERNAL_SOURCE_ROLE, INTERNAL_APPEND_ORDER],
        )

    def test_collect_target_columns_ignores_legacy_internal_display_columns(self) -> None:
        old_source = SourceSelection(
            source_id="old",
            path=Path("old.xlsx"),
            sheet_name="Sheet1",
            dataset_role="old",
            columns=["邮箱", default_display_name(INTERNAL_SOURCE_FILE), default_display_name(INTERNAL_APPEND_ORDER)],
        )

        self.assertEqual(collect_target_columns({"old": old_source}), ["邮箱"])

    def test_partial_manual_mapping_does_not_auto_fill_other_columns(self) -> None:
        old_source = SourceSelection(
            source_id="old",
            path=Path("old.xlsx"),
            sheet_name="Sheet1",
            dataset_role="old",
            columns=["邮箱", "密码"],
            row_count=0,
        )
        new_source = SourceSelection(
            source_id="new",
            path=Path("new.xlsx"),
            sheet_name="Sheet1",
            dataset_role="new",
            columns=["列1", "列2"],
            row_count=1,
            source_column_mapping={"列1": "邮箱", "列2": ""},
            mapping_confirmed=True,
        )
        cache = {
            "old": pd.DataFrame(columns=["邮箱", "密码"]),
            "new": pd.DataFrame([{"列1": "a@example.com", "列2": "pw123456"}]),
        }

        combined = combine_enabled_sources({"old": old_source, "new": new_source}, cache)
        result = process_dataframe(combined, PipelineConfig(include_source_columns=False))

        self.assertEqual(result.dataframe["邮箱"].tolist(), ["a@example.com"])
        self.assertTrue(pd.isna(result.dataframe.iloc[0]["密码"]))

    def test_contains_filter_treats_value_as_literal_text(self) -> None:
        series = pd.Series(["a[b", "abc", "."])

        contains_bracket = build_filter_mask(series, "contains", "[")
        contains_dot = build_filter_mask(series, "contains", ".")

        self.assertEqual(contains_bracket.tolist(), [True, False, False])
        self.assertEqual(contains_dot.tolist(), [False, False, True])

    def test_old_rows_stay_ahead_of_new_only_rows_regardless_of_import_order(self) -> None:
        old_source = SourceSelection(
            source_id="old",
            path=Path("old.xlsx"),
            sheet_name="Sheet1",
            dataset_role="old",
            columns=["email", "password"],
            row_count=3,
        )
        new_source = SourceSelection(
            source_id="new",
            path=Path("new.xlsx"),
            sheet_name="Sheet1",
            dataset_role="new",
            columns=["email", "password"],
            row_count=2,
            source_column_mapping={"email": "email", "password": "password"},
            mapping_confirmed=True,
        )
        cache = {
            "new": pd.DataFrame(
                [
                    {"email": "b@example.com", "password": "new-b"},
                    {"email": "d@example.com", "password": "new-d"},
                ]
            ),
            "old": pd.DataFrame(
                [
                    {"email": "a@example.com", "password": "old-a"},
                    {"email": "b@example.com", "password": "old-b"},
                    {"email": "c@example.com", "password": "old-c"},
                ]
            ),
        }

        combined = combine_enabled_sources({"new": new_source, "old": old_source}, cache)
        result = process_dataframe(
            combined,
            PipelineConfig(
                duplicate_keys=["email"],
                duplicate_strategy="update_and_append",
                include_source_columns=False,
            ),
        )

        self.assertEqual(
            result.dataframe["email"].tolist(),
            ["a@example.com", "b@example.com", "c@example.com", "d@example.com"],
        )

    def test_update_and_append_keeps_existing_value_when_new_cell_is_empty(self) -> None:
        old_source = SourceSelection(
            source_id="old",
            path=Path("old.xlsx"),
            sheet_name="Sheet1",
            dataset_role="old",
            columns=["email", "password", "phone", "note"],
            row_count=1,
        )
        new_source = SourceSelection(
            source_id="new",
            path=Path("new.xlsx"),
            sheet_name="Sheet1",
            dataset_role="new",
            columns=["email", "password", "phone"],
            row_count=2,
            source_column_mapping={"email": "email", "password": "password", "phone": "phone"},
            mapping_confirmed=True,
        )
        cache = {
            "old": pd.DataFrame(
                [
                    {
                        "email": "a@example.com",
                        "password": "old-pass",
                        "phone": "0911000000",
                        "note": "keep-note",
                    }
                ]
            ),
            "new": pd.DataFrame(
                [
                    {"email": "a@example.com", "password": "new-pass", "phone": pd.NA},
                    {"email": "c@example.com", "password": "c-pass", "phone": "0933000000"},
                ]
            ),
        }

        combined = combine_enabled_sources({"old": old_source, "new": new_source}, cache)
        result = process_dataframe(
            combined,
            PipelineConfig(
                duplicate_keys=["email"],
                duplicate_strategy="update_and_append",
                include_source_columns=False,
            ),
        )

        self.assertEqual(result.dataframe["email"].tolist(), ["a@example.com", "c@example.com"])
        row_a = result.dataframe[result.dataframe["email"] == "a@example.com"].iloc[0]
        row_c = result.dataframe[result.dataframe["email"] == "c@example.com"].iloc[0]
        self.assertEqual(row_a["password"], "new-pass")
        self.assertEqual(row_a["phone"], "0911000000")
        self.assertEqual(row_a["note"], "keep-note")
        self.assertEqual(row_c["password"], "c-pass")
        self.assertEqual(row_c["phone"], "0933000000")
        self.assertTrue(pd.isna(row_c["note"]))

    def test_update_and_append_drops_new_rows_with_blank_primary_keys(self) -> None:
        old_source = SourceSelection(
            source_id="old",
            path=Path("old.xlsx"),
            sheet_name="Sheet1",
            dataset_role="old",
            columns=["email", "note"],
            row_count=2,
        )
        new_source = SourceSelection(
            source_id="new",
            path=Path("new.xlsx"),
            sheet_name="Sheet1",
            dataset_role="new",
            columns=["email", "note"],
            row_count=2,
            source_column_mapping={"email": "email", "note": "note"},
            mapping_confirmed=True,
        )
        cache = {
            "old": pd.DataFrame(
                [
                    {"email": "a@example.com", "note": "keep-a"},
                    {"email": pd.NA, "note": "keep-old-blank"},
                ]
            ),
            "new": pd.DataFrame(
                [
                    {"email": "b@example.com", "note": "append-b"},
                    {"email": pd.NA, "note": "should-drop"},
                ]
            ),
        }

        combined = combine_enabled_sources({"old": old_source, "new": new_source}, cache)
        result = process_dataframe(
            combined,
            PipelineConfig(
                duplicate_keys=["email"],
                duplicate_strategy="update_and_append",
                include_source_columns=False,
            ),
        )

        self.assertEqual(result.dataframe["email"].tolist(), ["a@example.com", pd.NA, "b@example.com"])
        self.assertEqual(result.dataframe["note"].tolist(), ["keep-a", "keep-old-blank", "append-b"])

    def test_update_and_append_matches_when_any_selected_key_matches(self) -> None:
        dataframe = pd.DataFrame(
            [
                {
                    "email": "a@example.com",
                    "phone": "0911000000",
                    "password": "old-pass",
                    "note": "old-note",
                    "__source_file": "old.xlsx",
                    "__source_sheet": "Sheet1",
                    "__source_role": "old",
                    "__append_order": 1,
                },
                {
                    "email": "a@example.com",
                    "phone": pd.NA,
                    "password": pd.NA,
                    "note": "new-note",
                    "__source_file": "new.xlsx",
                    "__source_sheet": "Sheet1",
                    "__source_role": "new",
                    "__append_order": 2,
                },
            ]
        )

        result = process_dataframe(
            dataframe,
            PipelineConfig(
                duplicate_keys=["email", "phone"],
                duplicate_strategy="update_and_append",
                include_source_columns=False,
            ),
        )

        self.assertEqual(len(result.dataframe), 1)
        row = result.dataframe.iloc[0]
        self.assertEqual(row["email"], "a@example.com")
        self.assertEqual(row["phone"], "0911000000")
        self.assertEqual(row["password"], "old-pass")
        self.assertEqual(row["note"], "new-note")

    def test_update_and_append_keeps_new_duplicate_rows_when_no_old_row_matches(self) -> None:
        dataframe = pd.DataFrame(
            [
                {
                    "email": "dup@example.com",
                    "note": "first-row",
                    "password": pd.NA,
                    "__source_file": "new.xlsx",
                    "__source_sheet": "Sheet1",
                    "__source_role": "new",
                    "__append_order": 1,
                },
                {
                    "email": "dup@example.com",
                    "note": pd.NA,
                    "password": "second-pass",
                    "__source_file": "new.xlsx",
                    "__source_sheet": "Sheet1",
                    "__source_role": "new",
                    "__append_order": 2,
                },
            ]
        )

        result = process_dataframe(
            dataframe,
            PipelineConfig(
                duplicate_keys=["email"],
                duplicate_strategy="update_and_append",
                include_source_columns=False,
            ),
        )

        self.assertEqual(result.writeback_dataframe["email"].tolist(), ["dup@example.com", "dup@example.com"])
        self.assertEqual(result.writeback_dataframe.iloc[0]["note"], "first-row")
        self.assertTrue(pd.isna(result.writeback_dataframe.iloc[1]["note"]))
        self.assertTrue(pd.isna(result.writeback_dataframe.iloc[0]["password"]))
        self.assertEqual(result.writeback_dataframe.iloc[1]["password"], "second-pass")

    def test_update_only_skips_missing_primary_keys(self) -> None:
        old_source = SourceSelection(
            source_id="old",
            path=Path("old.xlsx"),
            sheet_name="Sheet1",
            dataset_role="old",
            columns=["email", "password", "phone", "note"],
            row_count=1,
        )
        new_source = SourceSelection(
            source_id="new",
            path=Path("new.xlsx"),
            sheet_name="Sheet1",
            dataset_role="new",
            columns=["email", "password", "phone"],
            row_count=2,
            source_column_mapping={"email": "email", "password": "password", "phone": "phone"},
            mapping_confirmed=True,
        )
        cache = {
            "old": pd.DataFrame(
                [
                    {
                        "email": "a@example.com",
                        "password": "old-pass",
                        "phone": "0911000000",
                        "note": "keep-note",
                    }
                ]
            ),
            "new": pd.DataFrame(
                [
                    {"email": "a@example.com", "password": "new-pass", "phone": pd.NA},
                    {"email": "c@example.com", "password": "c-pass", "phone": "0933000000"},
                ]
            ),
        }

        combined = combine_enabled_sources({"old": old_source, "new": new_source}, cache)
        result = process_dataframe(
            combined,
            PipelineConfig(
                duplicate_keys=["email"],
                duplicate_strategy="update_only",
                include_source_columns=False,
            ),
        )

        self.assertEqual(result.dataframe["email"].tolist(), ["a@example.com"])
        row_a = result.dataframe.iloc[0]
        self.assertEqual(row_a["password"], "new-pass")
        self.assertEqual(row_a["phone"], "0911000000")
        self.assertEqual(row_a["note"], "keep-note")

    def test_update_only_drops_new_rows_with_blank_primary_keys(self) -> None:
        old_source = SourceSelection(
            source_id="old",
            path=Path("old.xlsx"),
            sheet_name="Sheet1",
            dataset_role="old",
            columns=["email", "password", "note"],
            row_count=1,
        )
        new_source = SourceSelection(
            source_id="new",
            path=Path("new.xlsx"),
            sheet_name="Sheet1",
            dataset_role="new",
            columns=["email", "password", "note"],
            row_count=2,
            source_column_mapping={"email": "email", "password": "password", "note": "note"},
            mapping_confirmed=True,
        )
        cache = {
            "old": pd.DataFrame(
                [
                    {
                        "email": "a@example.com",
                        "password": "old-pass",
                        "note": "keep-note",
                    }
                ]
            ),
            "new": pd.DataFrame(
                [
                    {"email": "a@example.com", "password": "new-pass", "note": "updated"},
                    {"email": pd.NA, "password": pd.NA, "note": "blank-key-row"},
                ]
            ),
        }

        combined = combine_enabled_sources({"old": old_source, "new": new_source}, cache)
        result = process_dataframe(
            combined,
            PipelineConfig(
                duplicate_keys=["email"],
                duplicate_strategy="update_only",
                include_source_columns=True,
            ),
        )

        self.assertEqual(len(result.dataframe), 1)
        self.assertEqual(result.dataframe.iloc[0]["email"], "a@example.com")
        self.assertEqual(result.dataframe.iloc[0]["password"], "new-pass")
        self.assertEqual(result.dataframe.iloc[0]["note"], "updated")

    def test_process_dataframe_keeps_writeback_dataframe_unaffected_by_column_settings(self) -> None:
        dataframe = pd.DataFrame(
            [
                {
                    "email": "a@example.com",
                    "password": "pass-1",
                    "__source_file": "old.xlsx",
                    "__source_sheet": "Sheet1",
                    "__source_role": "old",
                    "__append_order": 1,
                }
            ]
        )
        result = process_dataframe(
            dataframe,
            PipelineConfig(
                column_settings={
                    "email": ColumnSetting(visible=True, rename_to="邮箱"),
                    "password": ColumnSetting(visible=False, rename_to=""),
                },
                include_source_columns=False,
            ),
        )

        self.assertEqual(list(result.dataframe.columns), ["邮箱"])
        self.assertEqual(list(result.writeback_dataframe.columns), ["email", "password"])

    def test_write_dataframe_to_existing_excel_sheet_handles_pd_na(self) -> None:
        from openpyxl import Workbook, load_workbook

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "writeback.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Sheet1"
            worksheet["A1"] = "old"
            workbook.save(output_path)
            workbook.close()

            dataframe = pd.DataFrame(
                [
                    {"email": "a@example.com", "phone": pd.NA},
                    {"email": "b@example.com", "phone": "0911000000"},
                ]
            )
            write_dataframe_to_existing_excel_sheet(
                dataframe,
                output_path,
                "Sheet1",
                ExportSettings(output_format="xlsx"),
            )

            workbook = load_workbook(output_path)
            try:
                worksheet = workbook["Sheet1"]
                self.assertEqual(worksheet["A2"].value, "a@example.com")
                self.assertIsNone(worksheet["B2"].value)
                self.assertEqual(worksheet["B3"].value, "0911000000")
            finally:
                workbook.close()

    def test_write_dataframe_to_existing_excel_sheet_preserves_target_sheet_structure(self) -> None:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import PatternFill

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "styled-writeback.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Sheet1"
            worksheet["A1"] = "email"
            worksheet["A2"] = "old@example.com"
            worksheet["A2"].fill = PatternFill(fill_type="solid", fgColor="FF0000")
            worksheet["A3"] = "stale@example.com"
            worksheet.freeze_panes = "B2"
            worksheet.column_dimensions["A"].width = 50
            worksheet.merge_cells("C1:D1")
            workbook.save(output_path)
            workbook.close()

            dataframe = pd.DataFrame([{"email": "new@example.com"}])
            write_dataframe_to_existing_excel_sheet(
                dataframe,
                output_path,
                "Sheet1",
                ExportSettings(output_format="xlsx"),
            )

            workbook = load_workbook(output_path)
            try:
                worksheet = workbook["Sheet1"]
                self.assertEqual(worksheet.freeze_panes, "B2")
                self.assertEqual(worksheet.column_dimensions["A"].width, 50)
                self.assertEqual([str(item) for item in worksheet.merged_cells.ranges], ["C1:D1"])
                self.assertEqual(worksheet["A2"].fill.fill_type, "solid")
                self.assertEqual(worksheet["A2"].fill.fgColor.rgb, "00FF0000")
                self.assertEqual(worksheet["A2"].value, "new@example.com")
                self.assertIsNone(worksheet["A3"].value)
            finally:
                workbook.close()

    def test_export_full_workbook_rejects_overwriting_source_file(self) -> None:
        from openpyxl import Workbook, load_workbook

        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "old.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Sheet1"
            worksheet["A1"] = "email"
            worksheet["A2"] = "old@example.com"
            workbook.save(workbook_path)
            workbook.close()

            source = SourceSelection("old", workbook_path, "Sheet1", dataset_role="old")
            dataframe = pd.DataFrame([{"email": "new@example.com"}])

            with self.assertRaisesRegex(ValueError, "不能覆盖原文件"):
                export_dataframe_with_old_workbook(dataframe, source, workbook_path, ExportSettings(output_format="xlsx"))

            workbook = load_workbook(workbook_path)
            try:
                worksheet = workbook["Sheet1"]
                self.assertEqual(worksheet["A2"].value, "old@example.com")
            finally:
                workbook.close()


class ComparisonRegressionTests(unittest.TestCase):
    def test_canonical_internal_column_name_accepts_legacy_display_names(self) -> None:
        self.assertEqual(canonical_internal_column_name(default_display_name(INTERNAL_SOURCE_FILE)), INTERNAL_SOURCE_FILE)
        self.assertEqual(canonical_internal_column_name(default_display_name(INTERNAL_APPEND_ORDER)), INTERNAL_APPEND_ORDER)

    def test_preview_value_treats_pd_na_as_empty(self) -> None:
        self.assertEqual(preview_value(pd.NA), "")

    def test_preview_value_treats_newline_only_text_as_empty(self) -> None:
        self.assertEqual(preview_value("\n"), "")

    def test_preview_value_flattens_multiline_text_for_single_line_preview(self) -> None:
        self.assertEqual(preview_value("line1\r\nline2\nline3"), "line1 line2 line3")

    def test_compare_row_values_treats_empty_and_pd_na_as_same(self) -> None:
        status, changed = compare_row_values({"email": None}, {"email": pd.NA}, ["email"])

        self.assertEqual(status, "same")
        self.assertEqual(changed, set())

    def test_compare_row_values_ignores_internal_compare_columns(self) -> None:
        status, changed = compare_row_values(
            {"邮箱": "a@example.com", "导入顺序": 1},
            {"邮箱": "a@example.com", "导入顺序": 1906},
            ["邮箱", "导入顺序"],
            ignored_columns={"导入顺序"},
        )

        self.assertEqual(status, "same")
        self.assertEqual(changed, set())

    def test_get_ignored_compare_columns_tracks_visible_internal_columns_after_rename(self) -> None:
        ignored_columns = get_ignored_compare_columns(
            ["email", "__source_file", "__append_order"],
            PipelineConfig(
                include_source_columns=True,
                column_settings={
                    "__source_file": ColumnSetting(visible=True, rename_to="来源文件"),
                    "__append_order": ColumnSetting(visible=True, rename_to="导入顺序"),
                },
            ),
        )

        self.assertEqual(ignored_columns, {"来源文件", "导入顺序"})

    def test_get_ignored_compare_columns_accepts_legacy_internal_display_columns(self) -> None:
        ignored_columns = get_ignored_compare_columns(
            ["email", "来源文件", "导入顺序"],
            PipelineConfig(include_source_columns=True),
        )

        self.assertEqual(ignored_columns, {"来源文件", "导入顺序"})

    def test_compare_preview_keeps_duplicate_keys(self) -> None:
        before = pd.DataFrame([{"email": "a@example.com", "password": "old"}])
        after = pd.DataFrame(
            [
                {"email": "a@example.com", "password": "old"},
                {"email": "a@example.com", "password": "new"},
            ]
        )

        comparison = align_for_comparison(
            before,
            after,
            PipelineConfig(duplicate_keys=["email"]),
            {},
        )

        self.assertEqual(len(comparison.before_df), 2)
        self.assertEqual(len(comparison.after_df), 2)
        self.assertEqual(comparison.before_df["email"].tolist(), ["a@example.com", ""])
        self.assertEqual(comparison.after_df["email"].tolist(), ["a@example.com", "a@example.com"])
        self.assertEqual(comparison.statuses, ["same", "added"])
        self.assertEqual(comparison.changed_columns, [set(), {"email", "password"}])

    def test_compare_preview_ignores_internal_columns_when_marking_changes(self) -> None:
        before = pd.DataFrame([{"邮箱": "a@example.com", "导入顺序": 1.0}])
        after = pd.DataFrame([{"邮箱": "a@example.com", "导入顺序": 1906.0}])

        comparison = align_for_comparison(
            before,
            after,
            PipelineConfig(duplicate_keys=["邮箱"]),
            {},
            ignored_columns={"导入顺序"},
        )

        self.assertEqual(comparison.statuses, ["same"])
        self.assertEqual(comparison.changed_columns, [set()])

    def test_compare_preview_can_align_by_hidden_key_columns(self) -> None:
        before_display = pd.DataFrame([{"备注": "old-a"}, {"备注": "old-b"}])
        after_display = pd.DataFrame([{"备注": "old-a"}, {"备注": "new-c"}, {"备注": "old-b"}])
        before_key_df = pd.DataFrame([{"email": "a@example.com"}, {"email": "b@example.com"}])
        after_key_df = pd.DataFrame(
            [
                {"email": "a@example.com"},
                {"email": "c@example.com"},
                {"email": "b@example.com"},
            ]
        )

        comparison = align_for_comparison(
            before_display,
            after_display,
            PipelineConfig(duplicate_keys=["email"]),
            {"email": ColumnSetting(visible=False)},
            before_key_df=before_key_df,
            after_key_df=after_key_df,
            key_columns=["email"],
        )

        self.assertEqual(comparison.statuses, ["same", "added", "same"])
        self.assertEqual(comparison.before_df.to_dict(orient="records"), [{"备注": "old-a"}, {"备注": ""}, {"备注": "old-b"}])
        self.assertEqual(comparison.after_df.to_dict(orient="records"), [{"备注": "old-a"}, {"备注": "new-c"}, {"备注": "old-b"}])

    def test_compare_preview_does_not_explode_when_most_rows_have_blank_selected_key(self) -> None:
        before_display = pd.DataFrame(
            [
                {"备注": "same-1", "邮箱": "a@example.com"},
                {"备注": "same-2", "邮箱": "b@example.com"},
                {"备注": "old-mark", "邮箱": "c@example.com"},
            ]
        )
        after_display = pd.DataFrame(
            [
                {"备注": "same-1", "邮箱": "a@example.com"},
                {"备注": "same-2", "邮箱": "b@example.com"},
                {"备注": "new-mark", "邮箱": "c@example.com"},
            ]
        )
        before_key_df = pd.DataFrame([{"secret": ""}, {"secret": ""}, {"secret": "match-key"}])
        after_key_df = pd.DataFrame([{"secret": ""}, {"secret": ""}, {"secret": "match-key"}])

        comparison = align_for_comparison(
            before_display,
            after_display,
            PipelineConfig(duplicate_keys=["secret"]),
            {},
            before_key_df=before_key_df,
            after_key_df=after_key_df,
            key_columns=["secret"],
        )

        self.assertEqual(comparison.statuses, ["same", "same", "changed"])


if __name__ == "__main__":
    unittest.main()
