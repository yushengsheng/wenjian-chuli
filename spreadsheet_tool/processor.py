from __future__ import annotations

import os
import re
import shutil
from pathlib import Path
from typing import Iterable
from uuid import uuid4

import pandas as pd

from .models import (
    ColumnSetting,
    ExportSettings,
    FilterRule,
    PipelineConfig,
    ProcessResult,
    SourceSelection,
    UpdateRule,
)

INTERNAL_SOURCE_FILE = "__source_file"
INTERNAL_SOURCE_SHEET = "__source_sheet"
INTERNAL_SOURCE_ROLE = "__source_role"
INTERNAL_APPEND_ORDER = "__append_order"
INTERNAL_COLUMNS = [
    INTERNAL_SOURCE_FILE,
    INTERNAL_SOURCE_SHEET,
    INTERNAL_SOURCE_ROLE,
    INTERNAL_APPEND_ORDER,
]

DISPLAY_NAME_OVERRIDES = {
    INTERNAL_SOURCE_FILE: "来源文件",
    INTERNAL_SOURCE_SHEET: "来源工作表",
    INTERNAL_SOURCE_ROLE: "数据分组",
    INTERNAL_APPEND_ORDER: "导入顺序",
}

SUPPORTED_FILE_TYPES = {
    ".xlsx",
    ".xlsm",
    ".csv",
    ".tsv",
}

AUTO_WIDTH_SAMPLE_LIMIT = 2000

CSV_ENCODINGS = ("utf-8-sig", "utf-8", "gb18030", "gbk")
EMPTY_TEXT_VALUES = {"", "nan", "none", "null", "nat", "<na>"}
_INTEGERISH_TEXT = re.compile(r"^-?\d+\.0+$")
EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
HEX_ADDRESS_PATTERN = re.compile(r"^0x[a-fA-F0-9]{40}$")
UPPER_SECRET_PATTERN = re.compile(r"^[A-Z2-7]{16,32}$")
LOWER_SECRET_PATTERN = re.compile(r"^[a-z0-9]{16,32}$")
LONG_TOKEN_PATTERN = re.compile(r"^[A-Za-z0-9]{40,120}$")
SHORT_TOKEN_PATTERN = re.compile(r"^[A-Za-z0-9]{6,24}$")
CODE_PATTERN = re.compile(r"^[A-Za-z]{1,4}\d{1,6}$")
PHONE_PATTERN = re.compile(r"^\+?\d[\d\-\s()]{6,18}\d$")

HEADER_ALIASES = {
    "特殊备注": {"特殊备注", "备注", "remark", "note", "notes"},
    "邮箱": {"邮箱", "email", "mail", "e-mail", "outlook"},
    "邮箱密码": {"邮箱密码", "emailpassword", "mailpassword", "password", "邮箱pass", "mailpass"},
    "邮箱2fa": {"邮箱2fa", "email2fa", "mail2fa", "邮箱谷歌", "邮箱otp", "emailotp"},
    "币安2fa": {"币安2fa", "binance2fa", "币安谷歌", "币安otp", "binanceotp"},
    "币安充值地址": {"币安充值地址", "充值地址", "地址", "depositaddress", "walletaddress", "binanceaddress"},
    "apikey": {"apikey", "api_key", "api key", "key"},
    "apisecret": {"apisecret", "api_secret", "api secret", "secret"},
}

HEADER_ALIASES.update(
    {
        "\u7279\u6b8a\u5907\u6ce8": {
            "\u7279\u6b8a\u5907\u6ce8",
            "\u5907\u6ce8",
            "remark",
            "note",
            "notes",
        },
        "\u90ae\u7bb1": {
            "\u90ae\u7bb1",
            "email",
            "mail",
            "e-mail",
            "outlook",
        },
        "\u624b\u673a\u53f7": {
            "\u624b\u673a\u53f7",
            "\u624b\u673a",
            "\u7535\u8bdd",
            "\u8054\u7cfb\u7535\u8bdd",
            "phone",
            "mobile",
            "tel",
            "telephone",
        },
        "\u90ae\u7bb1\u5bc6\u7801": {
            "\u90ae\u7bb1\u5bc6\u7801",
            "emailpassword",
            "mailpassword",
            "password",
            "\u90ae\u7bb1pass",
            "mailpass",
        },
        "\u90ae\u7bb12fa": {
            "\u90ae\u7bb12fa",
            "\u90ae\u7bb12fa/Oauth",
            "\u90ae\u7bb12fa/oauth",
            "email2fa",
            "mail2fa",
            "\u90ae\u7bb1otp",
            "emailotp",
            "2fa",
            "oauth",
            "oath",
        },
        "\u5e01\u5b892fa": {
            "\u5e01\u5b892fa",
            "binance2fa",
            "\u5e01\u5b89otp",
            "binanceotp",
        },
        "\u5e01\u5b89\u5145\u503c\u5730\u5740": {
            "\u5e01\u5b89\u5145\u503c\u5730\u5740",
            "\u5145\u503c\u5730\u5740",
            "\u5730\u5740",
            "depositaddress",
            "walletaddress",
            "binanceaddress",
        },
    }
)

WEAK_KINDS = {"text", "unknown", "empty", "code", "number"}

FILTER_OPERATORS = {
    "equals": "等于",
    "not_equals": "不等于",
    "contains": "包含",
    "not_contains": "不包含",
    "greater_than": "大于",
    "greater_equal": "大于等于",
    "less_than": "小于",
    "less_equal": "小于等于",
    "is_empty": "为空",
    "not_empty": "不为空",
}

UPDATE_MODES = {
    "set_value": "整列赋值",
    "fill_empty": "空值补全",
    "replace_text": "文本替换",
    "replace_exact": "精确替换",
}

DUPLICATE_STRATEGIES = {
    "update_and_append": "更新并新增",
    "update_only": "仅更新不新增",
    "fill_old_empty": "仅用新数据补全老数据空值",
    "keep_first": "保留首条",
    "keep_last": "保留末条",
    "none": "不处理重复",
}

LEGACY_DUPLICATE_STRATEGY_ALIASES = {
    "new_overwrite_old": "update_and_append",
}

UNMAPPED_TARGET = object()


def default_visible(column_name: str, include_source_columns: bool = True) -> bool:
    if column_name in INTERNAL_COLUMNS:
        return include_source_columns
    return True


def default_display_name(column_name: str) -> str:
    return DISPLAY_NAME_OVERRIDES.get(column_name, column_name)


def load_sources_from_paths(
    paths: Iterable[str | Path],
) -> tuple[list[SourceSelection], dict[str, pd.DataFrame]]:
    sources: list[SourceSelection] = []
    cache: dict[str, pd.DataFrame] = {}

    for raw_path in paths:
        path = Path(raw_path)
        suffix = path.suffix.lower()
        if suffix not in SUPPORTED_FILE_TYPES:
            raise ValueError(f"暂不支持该文件类型: {path.name}")

        if suffix in {".csv", ".tsv"}:
            dataframe = read_delimited_file(path, "\t" if suffix == ".tsv" else ",")
            source_id = uuid4().hex
            sources.append(
                SourceSelection(
                    source_id=source_id,
                    path=path,
                    sheet_name="数据",
                    row_count=len(dataframe),
                    columns=list(dataframe.columns),
                )
            )
            cache[source_id] = dataframe
            continue

        with pd.ExcelFile(path, engine="openpyxl") as workbook:
            for sheet_name in workbook.sheet_names:
                raw = workbook.parse(
                    sheet_name=sheet_name,
                    dtype=object,
                    header=None,
                )
                dataframe = materialize_dataframe(raw)
                source_id = uuid4().hex
                sources.append(
                    SourceSelection(
                        source_id=source_id,
                        path=path,
                        sheet_name=sheet_name,
                        row_count=len(dataframe),
                        columns=list(dataframe.columns),
                    )
                )
                cache[source_id] = dataframe

    return sources, cache


def read_delimited_file(path: Path, separator: str) -> pd.DataFrame:
    last_error: Exception | None = None
    for encoding in CSV_ENCODINGS:
        try:
            raw = pd.read_csv(path, sep=separator, dtype=object, encoding=encoding, header=None)
            return materialize_dataframe(raw)
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error is not None:
        raise ValueError(f"无法识别文件编码: {path.name}") from last_error
    raw = pd.read_csv(path, sep=separator, dtype=object, header=None)
    return materialize_dataframe(raw)


def read_excel_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    raw = pd.read_excel(
        path,
        sheet_name=sheet_name,
        dtype=object,
        engine="openpyxl",
        header=None,
    )
    return materialize_dataframe(raw)


def materialize_dataframe(raw: pd.DataFrame) -> pd.DataFrame:
    if raw is None or raw.empty:
        return pd.DataFrame()

    prepared = raw.dropna(axis=0, how="all").reset_index(drop=True)
    if prepared.empty:
        return pd.DataFrame()

    use_header = first_row_looks_like_header(prepared)
    if use_header:
        header_values = prepared.iloc[0].tolist()
        body = prepared.iloc[1:].copy()
        selected_indexes: list[int] = []
        columns: list[str] = []
        for index, value in enumerate(header_values):
            column_series = body.iloc[:, index] if index < body.shape[1] else pd.Series(dtype=object)
            has_named_header = not is_empty_value(value)
            has_data = not is_empty_series(column_series).all() if not column_series.empty else False
            if has_named_header or has_data:
                selected_indexes.append(index)
                columns.append(normalize_column_name(value, index))
        dataframe = body.iloc[:, selected_indexes].copy() if selected_indexes else pd.DataFrame()
        dataframe.columns = columns
        return normalize_dataframe(dataframe, preserve_empty_columns=True)
    else:
        dataframe = prepared.copy()
        dataframe.columns = [f"列{index + 1}" for index in range(dataframe.shape[1])]
    return normalize_dataframe(dataframe)


def normalize_dataframe(dataframe: pd.DataFrame, preserve_empty_columns: bool = False) -> pd.DataFrame:
    if dataframe is None:
        return pd.DataFrame()

    normalized = dataframe.copy()
    normalized.columns = make_unique_column_names(
        [normalize_column_name(column_name, index) for index, column_name in enumerate(normalized.columns)]
    )
    normalized = normalized.dropna(axis=0, how="all")
    if not preserve_empty_columns:
        normalized = normalized.dropna(axis=1, how="all")
    normalized = normalized.reset_index(drop=True)
    return normalized


def normalize_column_name(column_name: object, index: int) -> str:
    if column_name is None:
        return f"列{index + 1}"
    text = str(column_name).strip()
    if text.lower() in EMPTY_TEXT_VALUES:
        return f"列{index + 1}"
    return text or f"列{index + 1}"


def make_unique_column_names(columns: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    unique_columns: list[str] = []
    for column in columns:
        count = seen.get(column, 0)
        if count == 0:
            unique_columns.append(column)
        else:
            unique_columns.append(f"{column}_{count + 1}")
        seen[column] = count + 1
    return unique_columns


def first_row_looks_like_header(raw: pd.DataFrame) -> bool:
    sample = raw.head(3).copy()
    if sample.empty:
        return True

    first_row = sample.iloc[0].tolist()
    second_row = sample.iloc[1].tolist() if len(sample) > 1 else []

    header_hits = sum(1 for value in first_row if header_alias_key(value))
    first_patterns = [infer_value_kind(value) for value in first_row]
    second_patterns = [infer_value_kind(value) for value in second_row]

    first_strong = sum(1 for kind in first_patterns if kind not in WEAK_KINDS)
    second_strong = sum(1 for kind in second_patterns if kind not in WEAK_KINDS)
    same_strong = sum(
        1
        for first_kind, second_kind in zip(first_patterns, second_patterns)
        if first_kind == second_kind and first_kind not in WEAK_KINDS
    )
    comparable_strong = min(first_strong, second_strong)

    if len(sample) == 1:
        if header_hits >= 1:
            return True
        if first_strong >= 1:
            return False
        return True

    if header_hits >= 1 and first_strong <= second_strong:
        return True
    if comparable_strong >= 1 and same_strong >= comparable_strong:
        return False
    if second_strong >= 3 and first_strong >= max(2, int(second_strong * 0.6)):
        return False
    if same_strong >= 3:
        return False
    if first_strong == 0 and second_strong >= 2:
        return True
    return True


def combine_enabled_sources(
    sources: dict[str, SourceSelection],
    cache: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    append_order = 1
    target_columns = collect_target_columns(sources)
    target_profiles = build_target_profiles(sources, cache, target_columns)

    enabled_sources = [source for source in sources.values() if source.enabled]
    # Keep old rows in their original template order and append new-only rows after them.
    ordered_sources = [source for source in enabled_sources if source.dataset_role == "old"]
    ordered_sources.extend(source for source in enabled_sources if source.dataset_role != "old")

    for source in ordered_sources:

        frame = cache[source.source_id].copy()
        if target_columns:
            if source.dataset_role == "new":
                frame = align_dataframe_to_target(
                    frame,
                    target_columns,
                    target_profiles,
                    source.source_column_mapping,
                )
            else:
                frame = frame.reindex(columns=target_columns)
        frame[INTERNAL_SOURCE_FILE] = source.path.name
        frame[INTERNAL_SOURCE_SHEET] = source.sheet_name
        frame[INTERNAL_SOURCE_ROLE] = source.dataset_role
        frame[INTERNAL_APPEND_ORDER] = range(append_order, append_order + len(frame))
        append_order += len(frame)
        frames.append(frame)

    if not frames:
        return pd.DataFrame()

    return pd.concat(frames, ignore_index=True, sort=False)


def collect_available_columns(
    sources: dict[str, SourceSelection],
    include_internal: bool = True,
) -> list[str]:
    old_columns = collect_target_columns(sources)
    if old_columns:
        columns = list(old_columns)
    else:
        columns = []
        seen: set[str] = set()
        for source in sources.values():
            if not source.enabled:
                continue
            for column in source.columns:
                if column in seen:
                    continue
                seen.add(column)
                columns.append(column)

    if include_internal:
        for column in INTERNAL_COLUMNS:
            if column not in columns:
                columns.append(column)

    return columns


def collect_target_columns(sources: dict[str, SourceSelection]) -> list[str]:
    columns: list[str] = []
    seen: set[str] = set()
    for source in sources.values():
        if not source.enabled or source.dataset_role != "old":
            continue
        for column in source.columns:
            if column in seen:
                continue
            seen.add(column)
            columns.append(column)
    return columns


def build_target_profiles(
    sources: dict[str, SourceSelection],
    cache: dict[str, pd.DataFrame],
    target_columns: list[str],
) -> dict[str, dict[str, object]]:
    profiles: dict[str, dict[str, object]] = {}
    for index, column in enumerate(target_columns):
        samples: list[object] = []
        for source in sources.values():
            if not source.enabled or source.dataset_role != "old":
                continue
            frame = cache[source.source_id]
            if column in frame.columns:
                series = frame[column].dropna().tolist()
                samples.extend(series[:20])
        profiles[column] = {
            "kind": infer_target_kind(column, samples),
            "index": index,
        }
    return profiles


def suggest_source_to_target_mapping(
    dataframe: pd.DataFrame,
    target_columns: list[str],
    target_profiles: dict[str, dict[str, object]],
) -> dict[str, str]:
    suggested = suggest_target_to_source_mapping(dataframe, target_columns, target_profiles)
    return {source_column: target_column for target_column, source_column in suggested.items()}


def build_direct_source_to_target_mapping(
    source_columns: Iterable[object],
    target_columns: list[str],
) -> dict[str, str]:
    direct_mapping = build_direct_mapping(source_columns, target_columns)
    return {source_column: target_column for target_column, source_column in direct_mapping.items()}


def is_direct_header_match_complete(
    source_columns: Iterable[object],
    target_columns: list[str],
) -> tuple[bool, dict[str, str]]:
    source_column_list = [str(column) for column in source_columns]
    direct_mapping = build_direct_source_to_target_mapping(source_column_list, target_columns)
    return len(direct_mapping) == len(source_column_list), direct_mapping


def suggest_target_to_source_mapping(
    dataframe: pd.DataFrame,
    target_columns: list[str],
    target_profiles: dict[str, dict[str, object]],
    excluded_sources: set[str] | None = None,
    excluded_targets: set[str] | None = None,
) -> dict[str, str]:
    excluded_sources = excluded_sources or set()
    excluded_targets = excluded_targets or set()

    mapping = build_direct_mapping(
        [column for column in dataframe.columns if column not in excluded_sources],
        [column for column in target_columns if column not in excluded_targets],
    )
    used_sources = set(mapping.values()) | set(excluded_sources)
    source_profiles = {
        column: {
            "kind": infer_source_column_kind(dataframe[column]),
            "index": index,
        }
        for index, column in enumerate(dataframe.columns)
        if column not in used_sources
    }

    for target_column in target_columns:
        if target_column in excluded_targets or target_column in mapping:
            continue
        target_kind = target_profiles.get(target_column, {}).get("kind", "unknown")
        candidate = choose_best_source_column(target_column, target_kind, source_profiles)
        if candidate is None:
            continue
        mapping[target_column] = candidate
        source_profiles.pop(candidate, None)
    return mapping


def align_dataframe_to_target(
    dataframe: pd.DataFrame,
    target_columns: list[str],
    target_profiles: dict[str, dict[str, object]],
    manual_source_mapping: dict[str, str] | None = None,
) -> pd.DataFrame:
    if dataframe.empty:
        return pd.DataFrame(columns=target_columns)

    aligned = pd.DataFrame(index=dataframe.index)
    used_sources: set[str] = set()
    assigned_targets: set[str] = set()

    if manual_source_mapping:
        for source_column, target_column in manual_source_mapping.items():
            if source_column not in dataframe.columns:
                continue
            if not target_column:
                used_sources.add(source_column)
                continue
            if target_column not in target_columns or target_column in assigned_targets:
                used_sources.add(source_column)
                continue
            aligned[target_column] = dataframe[source_column]
            used_sources.add(source_column)
            assigned_targets.add(target_column)

    suggested_mapping = suggest_target_to_source_mapping(
        dataframe,
        target_columns,
        target_profiles,
        excluded_sources=used_sources,
        excluded_targets=assigned_targets,
    )
    for target_column, source_column in suggested_mapping.items():
        aligned[target_column] = dataframe[source_column]
        used_sources.add(source_column)
        assigned_targets.add(target_column)

    for target_column in target_columns:
        if target_column not in aligned.columns:
            aligned[target_column] = UNMAPPED_TARGET

    return aligned.reindex(columns=target_columns)


def build_direct_mapping(source_columns: Iterable[object], target_columns: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    used: set[str] = set()
    target_by_normalized_name: dict[str, str] = {}
    for column in target_columns:
        normalized_name = normalize_header_text(str(column))
        if normalized_name and normalized_name not in target_by_normalized_name:
            target_by_normalized_name[normalized_name] = column

    for source_column in source_columns:
        source_text = str(source_column)
        normalized_name = normalize_header_text(source_text)
        if not normalized_name or normalized_name not in target_by_normalized_name:
            continue
        target_column = target_by_normalized_name[normalized_name]
        if target_column in mapping or source_text in used:
            continue
        mapping[target_column] = source_text
        used.add(source_text)

    target_by_alias = {
        alias: column
        for column in target_columns
        for alias in [header_alias_key(column)]
        if alias and column not in mapping
    }
    for source_column in source_columns:
        source_text = str(source_column)
        if source_text in used:
            continue
        alias = header_alias_key(source_column)
        if not alias or alias not in target_by_alias:
            continue
        target_column = target_by_alias[alias]
        if target_column in mapping or source_text in used:
            continue
        mapping[target_column] = source_text
        used.add(source_text)
    return mapping


def choose_best_source_column(
    target_column: str,
    target_kind: str,
    source_profiles: dict[str, dict[str, object]],
) -> str | None:
    if target_kind in WEAK_KINDS:
        return None

    candidates: list[tuple[int, int, str]] = []
    for source_column, profile in source_profiles.items():
        source_kind = profile["kind"]
        if not kinds_are_compatible(target_kind, source_kind):
            continue
        penalty = abs(int(profile["index"]))
        bonus = 0 if source_kind == target_kind else 1
        candidates.append((bonus, penalty, source_column))

    if not candidates:
        return None

    candidates.sort()
    return candidates[0][2]


def kinds_are_compatible(target_kind: str, source_kind: str) -> bool:
    if target_kind == source_kind:
        return True
    compatible_groups = [
        {"long_token"},
        {"wallet_address"},
        {"email"},
        {"phone"},
        {"otp_lower"},
        {"otp_or_oauth", "otp_lower", "long_token"},
        {"otp_upper"},
        {"short_token", "password_like"},
    ]
    for group in compatible_groups:
        if target_kind in group and source_kind in group:
            return True
    return False


def infer_target_kind(column_name: str, samples: list[object]) -> str:
    alias = header_alias_key(column_name)
    normalized_name = normalize_header_text(str(column_name))
    if alias == "\u90ae\u7bb1":
        return "email"
    if alias == "\u624b\u673a\u53f7":
        return "phone"
    if alias == "\u90ae\u7bb1\u5bc6\u7801":
        return "password_like"
    if alias == "\u90ae\u7bb12fa":
        if "oauth" in normalized_name or "oath" in normalized_name:
            return "otp_or_oauth"
        return "otp_lower"
    if alias == "\u5e01\u5b892fa":
        return "otp_upper"
    if alias == "\u5e01\u5b89\u5145\u503c\u5730\u5740":
        return "wallet_address"
    if alias == "\u7279\u6b8a\u5907\u6ce8":
        return "text"
    if alias == "邮箱":
        return "email"
    if alias == "邮箱密码":
        return "password_like"
    if alias == "邮箱2fa":
        return "otp_lower"
    if alias == "币安2fa":
        return "otp_upper"
    if alias == "币安充值地址":
        return "wallet_address"
    if alias in {"apikey", "apisecret"}:
        return "long_token"
    if alias == "特殊备注":
        return "text"
    return infer_values_kind(samples)


def infer_source_column_kind(series: pd.Series) -> str:
    values = series.dropna().tolist()[:30]
    return infer_values_kind(values)


def infer_values_kind(values: list[object]) -> str:
    counts: dict[str, int] = {}
    for value in values:
        kind = infer_value_kind(value)
        counts[kind] = counts.get(kind, 0) + 1
    if not counts:
        return "empty"
    return max(counts.items(), key=lambda item: item[1])[0]


def infer_value_kind(value: object) -> str:
    if is_empty_value(value):
        return "empty"
    text = str(value).strip()
    if EMAIL_PATTERN.fullmatch(text):
        return "email"
    if looks_like_phone(text):
        return "phone"
    if HEX_ADDRESS_PATTERN.fullmatch(text):
        return "wallet_address"
    if UPPER_SECRET_PATTERN.fullmatch(text):
        return "otp_upper"
    if LOWER_SECRET_PATTERN.fullmatch(text):
        return "otp_lower"
    if LONG_TOKEN_PATTERN.fullmatch(text):
        return "long_token"
    if SHORT_TOKEN_PATTERN.fullmatch(text):
        return "password_like"
    if CODE_PATTERN.fullmatch(text):
        return "code"
    if looks_like_number(text):
        return "number"
    return "text"


def looks_like_number(text: str) -> bool:
    try:
        float(text)
        return True
    except ValueError:
        return False


def looks_like_phone(text: str) -> bool:
    digits = re.sub(r"\D", "", text)
    if len(digits) < 7 or len(digits) > 15:
        return False
    if EMAIL_PATTERN.fullmatch(text) or HEX_ADDRESS_PATTERN.fullmatch(text):
        return False
    if PHONE_PATTERN.fullmatch(text):
        return True
    return text.isdigit() and len(digits) >= 8


def header_alias_key(value: object) -> str:
    if value is None:
        return ""
    normalized = normalize_header_text(str(value))
    for target, aliases in HEADER_ALIASES.items():
        if normalized in {normalize_header_text(alias) for alias in aliases | {target}}:
            return target
    return ""


def normalize_header_text(text: str) -> str:
    return re.sub(r"[\s_\-:/\\]+", "", text.strip().lower())


def process_dataframe(dataframe: pd.DataFrame, config: PipelineConfig) -> ProcessResult:
    working = dataframe.copy()
    summary_lines = [f"初始记录数: {len(working)}"]

    if working.empty:
        writeback_output = build_writeback_dataframe(working)
        output = apply_column_settings(working, config)
        summary_lines.append("没有可处理的数据。")
        return ProcessResult(dataframe=output, writeback_dataframe=writeback_output, summary_lines=summary_lines)

    if config.duplicate_keys and config.duplicate_strategy != "none":
        before = len(working)
        working = apply_duplicate_strategy(working, config.duplicate_keys, config.duplicate_strategy)
        summary_lines.append(f"主键合并后: {len(working)} 行，减少 {before - len(working)} 行")
    else:
        summary_lines.append("主键合并: 未启用")

    working = cleanup_unmapped_targets(working)

    if config.update_rules:
        working = apply_update_rules(working, config.update_rules)
        summary_lines.append(f"更新规则数: {len(config.update_rules)}")
    else:
        summary_lines.append("更新规则数: 0")

    if config.filter_rules:
        before = len(working)
        working = apply_filter_rules(working, config.filter_rules)
        summary_lines.append(f"筛选后: {len(working)} 行，过滤 {before - len(working)} 行")
    else:
        summary_lines.append("筛选规则数: 0")

    writeback_output = build_writeback_dataframe(working)
    working = apply_column_settings(working, config)
    summary_lines.append(f"输出列数: {len(working.columns)}")
    return ProcessResult(dataframe=working, writeback_dataframe=writeback_output, summary_lines=summary_lines)


def apply_duplicate_strategy(
    dataframe: pd.DataFrame,
    keys: list[str],
    strategy: str,
) -> pd.DataFrame:
    strategy = normalize_duplicate_strategy(strategy)
    valid_keys = [key for key in keys if key in dataframe.columns]
    if not valid_keys:
        return dataframe.reset_index(drop=True)

    has_key_mask = pd.Series(False, index=dataframe.index)
    for key in valid_keys:
        has_key_mask = has_key_mask | ~is_empty_series(dataframe[key])

    keyed_rows = dataframe[has_key_mask].copy()
    blank_rows = dataframe[~has_key_mask].copy()
    if keyed_rows.empty:
        return dataframe.reset_index(drop=True)

    keyed_rows = keyed_rows.sort_values(INTERNAL_APPEND_ORDER, kind="mergesort")
    helper_columns = attach_normalized_key_columns(keyed_rows, valid_keys)

    if strategy == "keep_first":
        deduped = keyed_rows.drop_duplicates(subset=helper_columns, keep="first")
    elif strategy == "keep_last":
        deduped = keyed_rows.drop_duplicates(subset=helper_columns, keep="last")
    elif strategy in {"update_and_append", "update_only", "fill_old_empty"}:
        deduped = merge_keyed_rows_by_role(keyed_rows, helper_columns, valid_keys, strategy)
    else:
        deduped = keyed_rows

    deduped = deduped.drop(columns=helper_columns, errors="ignore")
    merged = pd.concat([deduped, blank_rows], ignore_index=True, sort=False)
    merged = merged.sort_values(INTERNAL_APPEND_ORDER, kind="mergesort").reset_index(drop=True)
    return merged


def attach_normalized_key_columns(dataframe: pd.DataFrame, keys: list[str]) -> list[str]:
    helper_columns: list[str] = []
    for index, key in enumerate(keys):
        helper_column = f"__group_key_{index}"
        dataframe[helper_column] = dataframe[key].map(normalize_key_value)
        helper_columns.append(helper_column)
    return helper_columns


def normalize_key_value(value: object) -> str:
    if is_empty_value(value):
        return ""

    if isinstance(value, bool):
        return str(value)

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if pd.isna(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return format(value, "g")

    text = str(value).strip()
    lowered = text.lower()
    if lowered in EMPTY_TEXT_VALUES:
        return ""
    if _INTEGERISH_TEXT.fullmatch(text):
        return text.split(".", 1)[0]
    return text


def merge_keyed_rows_by_role(
    dataframe: pd.DataFrame,
    helper_columns: list[str],
    key_columns: list[str],
    strategy: str,
) -> pd.DataFrame:
    strategy = normalize_duplicate_strategy(strategy)
    output_columns = [column for column in dataframe.columns if column not in helper_columns]
    rows: list[dict[str, object]] = []

    grouped = dataframe.groupby(helper_columns, dropna=False, sort=False)
    for _, group in grouped:
        old_group = group[group[INTERNAL_SOURCE_ROLE] == "old"].sort_values(INTERNAL_APPEND_ORDER, kind="mergesort")
        new_group = group[group[INTERNAL_SOURCE_ROLE] == "new"].sort_values(INTERNAL_APPEND_ORDER, kind="mergesort")

        preserve_full_row = strategy in {"update_and_append", "update_only"}
        old_row = merge_rows_within_role(old_group, output_columns, key_columns, preserve_full_row=preserve_full_row)
        new_row = merge_rows_within_role(new_group, output_columns, key_columns, preserve_full_row=preserve_full_row)

        if old_row and new_row:
            merged_row = overlay_new_row(old_row, new_row, key_columns, strategy)
        elif old_row:
            merged_row = old_row
        elif strategy == "update_only":
            continue
        else:
            merged_row = new_row

        merged_row[INTERNAL_SOURCE_FILE] = join_distinct_non_empty(group[INTERNAL_SOURCE_FILE])
        merged_row[INTERNAL_SOURCE_SHEET] = join_distinct_non_empty(group[INTERNAL_SOURCE_SHEET])
        merged_row[INTERNAL_SOURCE_ROLE] = join_distinct_non_empty(group[INTERNAL_SOURCE_ROLE])
        if old_row:
            merged_row[INTERNAL_APPEND_ORDER] = old_row.get(INTERNAL_APPEND_ORDER)
        elif new_row:
            merged_row[INTERNAL_APPEND_ORDER] = new_row.get(INTERNAL_APPEND_ORDER)
        else:
            merged_row[INTERNAL_APPEND_ORDER] = group[INTERNAL_APPEND_ORDER].max()
        rows.append(merged_row)

    result = pd.DataFrame(rows, columns=output_columns)
    return result.sort_values(INTERNAL_APPEND_ORDER, kind="mergesort").reset_index(drop=True)


def merge_rows_within_role(
    group: pd.DataFrame,
    output_columns: list[str],
    key_columns: list[str],
    preserve_full_row: bool = False,
) -> dict[str, object]:
    if group.empty:
        return {}

    if preserve_full_row:
        row_data = {column: group.iloc[-1][column] for column in output_columns}
        for column in key_columns:
            row_data[column] = best_key_display_value(group[column])
        for column in {INTERNAL_SOURCE_FILE, INTERNAL_SOURCE_SHEET, INTERNAL_SOURCE_ROLE}:
            if column in row_data:
                row_data[column] = join_distinct_non_empty(group[column])
        if INTERNAL_APPEND_ORDER in row_data:
            row_data[INTERNAL_APPEND_ORDER] = group[INTERNAL_APPEND_ORDER].iloc[-1]
        return row_data

    row_data: dict[str, object] = {}
    for column in output_columns:
        series = group[column]
        if column == INTERNAL_APPEND_ORDER:
            row_data[column] = series.iloc[-1]
        elif column in key_columns:
            row_data[column] = best_key_display_value(series)
        elif column in {INTERNAL_SOURCE_FILE, INTERNAL_SOURCE_SHEET, INTERNAL_SOURCE_ROLE}:
            row_data[column] = join_distinct_non_empty(series)
        else:
            row_data[column] = last_non_empty(series)
    return row_data


def overlay_new_row(
    old_row: dict[str, object],
    new_row: dict[str, object],
    key_columns: list[str],
    strategy: str,
) -> dict[str, object]:
    strategy = normalize_duplicate_strategy(strategy)
    merged_row = dict(old_row)
    for column, new_value in new_row.items():
        if column == INTERNAL_APPEND_ORDER:
            merged_row[column] = max_value(merged_row.get(column), new_value)
            continue
        if column in {INTERNAL_SOURCE_FILE, INTERNAL_SOURCE_SHEET, INTERNAL_SOURCE_ROLE}:
            merged_row[column] = join_distinct_values([merged_row.get(column), new_value])
            continue
        if is_unmapped_target(new_value):
            continue
        if column in key_columns:
            if not is_empty_value(new_value):
                merged_row[column] = normalize_key_value(new_value)
            continue

        if strategy in {"update_and_append", "update_only"}:
            if values_differ(merged_row.get(column), new_value):
                merged_row[column] = new_value
        elif strategy == "fill_old_empty":
            if is_empty_value(merged_row.get(column)) and not is_empty_value(new_value):
                merged_row[column] = new_value

    return merged_row


def max_value(left: object, right: object) -> object:
    if left is None:
        return right
    if right is None:
        return left
    return max(left, right)


def best_key_display_value(series: pd.Series) -> str | object:
    for value in reversed(series.tolist()):
        normalized = normalize_key_value(value)
        if normalized != "":
            return normalized
    return last_non_empty(series)


def join_distinct_non_empty(series: pd.Series) -> str:
    return join_distinct_values(series.tolist())


def join_distinct_values(values: Iterable[object]) -> str:
    seen: set[str] = set()
    output: list[str] = []
    for value in values:
        if is_empty_value(value):
            continue
        text = str(value).strip()
        if text in seen:
            continue
        seen.add(text)
        output.append(text)
    return " | ".join(output)


def last_non_empty(series: pd.Series) -> object:
    for value in reversed(series.tolist()):
        if not is_empty_value(value):
            return value
    return series.iloc[-1] if not series.empty else None


def apply_filter_rules(dataframe: pd.DataFrame, rules: list[FilterRule]) -> pd.DataFrame:
    filtered = dataframe.copy()
    for rule in rules:
        if rule.column not in filtered.columns:
            continue
        mask = build_filter_mask(filtered[rule.column], rule.operator, rule.value)
        filtered = filtered[mask].copy()
    return filtered.reset_index(drop=True)


def build_filter_mask(series: pd.Series, operator: str, value: str) -> pd.Series:
    text_series = series.fillna("").astype(str)

    if operator == "equals":
        return text_series == value
    if operator == "not_equals":
        return text_series != value
    if operator == "contains":
        return text_series.str.contains(value, na=False, regex=False)
    if operator == "not_contains":
        return ~text_series.str.contains(value, na=False, regex=False)
    if operator in {"greater_than", "greater_equal", "less_than", "less_equal"}:
        numeric_series = pd.to_numeric(series, errors="coerce")
        numeric_value = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
        if pd.isna(numeric_value):
            raise ValueError(f"筛选值不是有效数字: {value}")
        if operator == "greater_than":
            return numeric_series > numeric_value
        if operator == "greater_equal":
            return numeric_series >= numeric_value
        if operator == "less_than":
            return numeric_series < numeric_value
        return numeric_series <= numeric_value
    if operator == "is_empty":
        return is_empty_series(series)
    if operator == "not_empty":
        return ~is_empty_series(series)

    raise ValueError(f"未知筛选条件: {operator}")


def apply_update_rules(dataframe: pd.DataFrame, rules: list[UpdateRule]) -> pd.DataFrame:
    updated = dataframe.copy()
    for rule in rules:
        if rule.column not in updated.columns:
            continue

        if rule.mode == "set_value":
            updated.loc[:, rule.column] = rule.replace_value
            continue

        if rule.mode == "fill_empty":
            empty_mask = is_empty_series(updated[rule.column])
            updated.loc[empty_mask, rule.column] = rule.replace_value
            continue

        if rule.mode == "replace_text":
            if not rule.find_value:
                continue
            updated.loc[:, rule.column] = updated[rule.column].map(
                lambda value: replace_text_value(value, rule.find_value, rule.replace_value)
            )
            continue

        if rule.mode == "replace_exact":
            text_series = updated[rule.column].fillna("").astype(str)
            updated.loc[text_series == rule.find_value, rule.column] = rule.replace_value

    return updated


def replace_text_value(value: object, source: str, target: str) -> object:
    if is_empty_value(value):
        return value
    return str(value).replace(source, target)


def apply_column_settings(dataframe: pd.DataFrame, config: PipelineConfig) -> pd.DataFrame:
    selected_columns: list[str] = []
    rename_map: dict[str, str] = {}

    for column in dataframe.columns:
        setting = config.column_settings.get(column, ColumnSetting())
        visible = setting.visible
        if column in INTERNAL_COLUMNS and not config.include_source_columns:
            visible = False
        if not visible:
            continue
        selected_columns.append(column)
        rename_map[column] = setting.rename_to.strip() or default_display_name(column)

    output = dataframe.loc[:, selected_columns].copy()
    output = output.rename(columns=make_unique_rename_map(rename_map))
    return output.reset_index(drop=True)


def make_unique_rename_map(rename_map: dict[str, str]) -> dict[str, str]:
    seen: dict[str, int] = {}
    unique_map: dict[str, str] = {}
    for original, target in rename_map.items():
        count = seen.get(target, 0)
        if count == 0:
            unique_map[original] = target
        else:
            unique_map[original] = f"{target}_{count + 1}"
        seen[target] = count + 1
    return unique_map


def export_dataframe(
    dataframe: pd.DataFrame,
    output_path: str | Path,
    settings: ExportSettings,
) -> None:
    path = Path(output_path)
    if settings.output_format == "csv":
        dataframe.to_csv(path, index=False, encoding="utf-8-sig")
        return
    export_to_excel(dataframe, path, settings)


def export_dataframe_with_old_workbook(
    dataframe: pd.DataFrame,
    source: SourceSelection,
    output_path: str | Path,
    settings: ExportSettings,
) -> None:
    source_path = source.path
    if source_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("仅支持基于 Excel 老文件导出整本工作簿。")

    target_path = Path(output_path)
    if paths_refer_to_same_file(source_path, target_path):
        raise ValueError("完整老文件导出不能覆盖原文件，请选择新的保存位置。")

    shutil.copy2(source_path, target_path)

    write_dataframe_to_existing_excel_sheet(dataframe, target_path, source.sheet_name, settings)


def export_to_excel(
    dataframe: pd.DataFrame,
    output_path: Path,
    settings: ExportSettings,
) -> None:
    excel_ready = prepare_dataframe_for_excel(dataframe)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        excel_ready.to_excel(writer, index=False, sheet_name=settings.sheet_name)
        worksheet = writer.book[settings.sheet_name]
        if settings.freeze_header:
            worksheet.freeze_panes = "A2"
        if settings.style_header:
            apply_header_style(worksheet)
        if settings.auto_width:
            apply_auto_width(worksheet, dataframe)


def write_dataframe_back_to_source(
    dataframe: pd.DataFrame,
    source: SourceSelection,
    settings: ExportSettings,
) -> None:
    path = source.path
    suffix = path.suffix.lower()

    if suffix == ".csv":
        dataframe.to_csv(path, index=False, encoding="utf-8-sig")
        return
    if suffix == ".tsv":
        dataframe.to_csv(path, index=False, sep="\t", encoding="utf-8-sig")
        return
    if suffix in {".xlsx", ".xlsm"}:
        write_dataframe_to_existing_excel_sheet(dataframe, path, source.sheet_name, settings)
        return

    raise ValueError(f"暂不支持回写到该文件类型: {path.name}")


def write_dataframe_to_existing_excel_sheet(
    dataframe: pd.DataFrame,
    output_path: Path,
    sheet_name: str,
    settings: ExportSettings,
) -> None:
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    workbook = load_workbook(output_path, keep_vba=output_path.suffix.lower() == ".xlsm")
    excel_ready = prepare_dataframe_for_excel(dataframe)
    try:
        sheet_index = len(workbook.sheetnames)
        if sheet_name in workbook.sheetnames:
            sheet_index = workbook.sheetnames.index(sheet_name)
            workbook.remove(workbook[sheet_name])

        worksheet = workbook.create_sheet(title=sheet_name, index=sheet_index)
        for row in dataframe_to_rows(excel_ready, index=False, header=True):
            worksheet.append(row)

        if settings.freeze_header:
            worksheet.freeze_panes = "A2"
        if settings.style_header and dataframe.shape[1] > 0:
            apply_header_style(worksheet)
        if settings.auto_width and dataframe.shape[1] > 0:
            apply_auto_width(worksheet, dataframe)
        workbook.save(output_path)
    finally:
        workbook.close()


def apply_header_style(worksheet: object) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    font = Font(bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


def apply_auto_width(worksheet: object, dataframe: pd.DataFrame | None = None) -> None:
    from openpyxl.utils import get_column_letter

    if dataframe is None:
        for column_cells in worksheet.columns:
            letter = column_cells[0].column_letter
            max_length = 0
            for cell in column_cells[:AUTO_WIDTH_SAMPLE_LIMIT]:
                text = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(text))
            worksheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 40)
        return

    sample = sample_auto_width_dataframe(dataframe)
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        max_length = len(str(column_name))
        if column_name in sample.columns:
            lengths = sample[column_name].map(cell_display_length)
            if not lengths.empty:
                max_length = max(max_length, int(lengths.max()))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 10), 40)


def sample_auto_width_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    if len(dataframe) <= AUTO_WIDTH_SAMPLE_LIMIT:
        return dataframe

    head_count = AUTO_WIDTH_SAMPLE_LIMIT // 2
    tail_count = AUTO_WIDTH_SAMPLE_LIMIT - head_count
    return pd.concat([dataframe.head(head_count), dataframe.tail(tail_count)], ignore_index=True)


def cell_display_length(value: object) -> int:
    if is_empty_value(value):
        return 0
    return len(str(value))


def is_empty_series(series: pd.Series) -> pd.Series:
    unmapped_mask = series.map(is_unmapped_target) if not series.empty else pd.Series(False, index=series.index)
    return unmapped_mask | series.isna() | series.fillna("").astype(str).str.strip().str.lower().isin(EMPTY_TEXT_VALUES)


def is_empty_value(value: object) -> bool:
    if is_unmapped_target(value):
        return True
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    text = str(value).strip()
    return text.lower() in EMPTY_TEXT_VALUES


def normalize_duplicate_strategy(strategy: str) -> str:
    return LEGACY_DUPLICATE_STRATEGY_ALIASES.get(strategy, strategy)


def is_unmapped_target(value: object) -> bool:
    return value is UNMAPPED_TARGET


def cleanup_unmapped_targets(dataframe: pd.DataFrame) -> pd.DataFrame:
    cleaned = dataframe.copy()
    for column in cleaned.columns:
        cleaned[column] = cleaned[column].map(lambda value: pd.NA if is_unmapped_target(value) else value)
    return cleaned


def build_writeback_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    writeback_columns = [column for column in dataframe.columns if column not in INTERNAL_COLUMNS]
    return dataframe.loc[:, writeback_columns].copy().reset_index(drop=True)


def prepare_dataframe_for_excel(dataframe: pd.DataFrame) -> pd.DataFrame:
    excel_ready = dataframe.copy().astype(object)
    return excel_ready.where(~excel_ready.isna(), None)


def paths_refer_to_same_file(left_path: str | Path, right_path: str | Path) -> bool:
    left = Path(left_path)
    right = Path(right_path)
    try:
        left_marker = os.path.normcase(str(left.resolve()))
        right_marker = os.path.normcase(str(right.resolve()))
    except OSError:
        left_marker = os.path.normcase(str(left))
        right_marker = os.path.normcase(str(right))
    return left_marker == right_marker


def values_differ(left: object, right: object) -> bool:
    return normalize_compare_value(left) != normalize_compare_value(right)


def normalize_compare_value(value: object) -> str:
    if is_unmapped_target(value) or is_empty_value(value):
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if pd.isna(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return format(value, "g")
    return str(value).strip()
